import os

from django.core.management import BaseCommand
from openpyxl import load_workbook

from TatSpeaks.settings import BASE_DIR
from backend.services.originals import create_original


class Command(BaseCommand):
    help = "Fill Originals to DB"

    def handle(self, *args, **options):
        wb = load_workbook(filename=os.path.join(BASE_DIR, 'originals.xlsx'))
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=2111, max_col=1):
            string = row[0].value
            if type(string) != str:
                continue
            if len(string) <= 255:
                create_original(string)
